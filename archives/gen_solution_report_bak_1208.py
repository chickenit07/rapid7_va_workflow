import pandas as pd
import lxml.etree as ET
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from tqdm import tqdm
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Report Directory
DOWNLOAD_PATH = os.getenv('DOWNLOAD_PATH', '')

def parse_xml_for_solutions_and_products(xml_file, vuln_ids):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    solutions = {}
    ip_product_map = {}
    ip_os_family_map = {}
    ip_hostname_map = {}  # New map for storing hostnames
    ip_office_product_map = {}  # Store Office product per IP
    ip_sharepoint_product_map = {}  # Store SharePoint product per IP
    ip_exchange_product_map = {}  # Store Exchange Server product per IP

    # Extract product names, OS family, and hostnames for each IP address
    for node in root.xpath(".//node"):
        ip_address = node.get("address")
        fingerprint = node.find(".//fingerprints/os")
        hostname_elem = node.find(".//names/name")  # Extract the hostname

        if fingerprint is not None:
            product_name = fingerprint.get("product")
            os_family = fingerprint.get("family")  # Extract the OS family
            ip_product_map[ip_address] = product_name
            ip_os_family_map[ip_address] = os_family

        if hostname_elem is not None:
            hostname = hostname_elem.text.strip()
            ip_hostname_map[ip_address] = hostname

        # Extract Office and SharePoint products from software fingerprints
        for sw_fp in node.findall('.//software/fingerprint'):
            product_value = sw_fp.get('product', '')
            if product_value:
                # Only capture Office entries that contain a year (like "Office 2013", "Office 2016")
                if 'Office' in product_value and any(char.isdigit() for char in product_value) and ip_address not in ip_office_product_map:
                    ip_office_product_map[ip_address] = product_value
                # Only capture SharePoint entries that contain a year (like "SharePoint 2016", "SharePoint 2019")
                if 'SharePoint' in product_value and any(char.isdigit() for char in product_value) and ip_address not in ip_sharepoint_product_map:
                    ip_sharepoint_product_map[ip_address] = product_value
                # Only capture Exchange Server entries that contain a year (like "Exchange Server 2016", "Exchange Server 2019")
                if 'Exchange Server' in product_value and any(char.isdigit() for char in product_value) and ip_address not in ip_exchange_product_map:
                    ip_exchange_product_map[ip_address] = product_value

    for vuln_id in tqdm(vuln_ids, desc="Parsing XML for Solutions"):
        solution_list = []

        # Check for solutions under ListItem
        xpath_query_list_item = f".//vulnerability[@id='{vuln_id}']/solution/ContainerBlockElement/UnorderedList/ListItem"
        list_items = root.xpath(xpath_query_list_item)

        for item in list_items:
            paragraphs = item.xpath(".//Paragraph")
            item_texts = []

            for i, para in enumerate(paragraphs):
                text_parts = []
                for elem in para.iter():
                    if elem.tag == "URLLink":
                        text_parts.append(elem.get("LinkURL"))
                    elif elem.text:
                        text_parts.append(elem.text.strip())
                paragraph_text = "".join(text_parts)

                if i == 0:
                    item_texts.append(paragraph_text)
                else:
                    item_texts.append(f"=> {paragraph_text}")

            full_item_text = " ".join(item_texts)
            solution_list.append(full_item_text)

        # Check for solutions directly under ContainerBlockElement/Paragraph
        xpath_query_paragraph = f".//vulnerability[@id='{vuln_id}']/solution/ContainerBlockElement/Paragraph"
        paragraphs = root.xpath(xpath_query_paragraph)

        for para in paragraphs:
            text_parts = []
            for elem in para.iter():
                if elem.tag == "URLLink":
                    text_parts.append(elem.get("LinkURL"))
                elif elem.text:
                    text_parts.append(elem.text.strip())
            paragraph_text = "".join(text_parts)
            if paragraph_text:
                solution_list.append(f"{vuln_id} => {paragraph_text}")

        # Check for solutions nested within ContainerBlockElement/Paragraph/Paragraph
        xpath_query_nested_paragraph = f".//vulnerability[@id='{vuln_id}']/solution/ContainerBlockElement/Paragraph/Paragraph"
        nested_paragraphs = root.xpath(xpath_query_nested_paragraph)

        for para in nested_paragraphs:
            text_parts = []
            for elem in para.iter():
                if elem.tag == "URLLink":
                    text_parts.append(elem.get("LinkURL"))
                elif elem.text:
                    text_parts.append(elem.text.strip())
            nested_text = "".join(text_parts)
            if nested_text:
                solution_list.append(f"{vuln_id} => {nested_text}")

        solutions[vuln_id] = solution_list

    # Print extracted Office, SharePoint, and Exchange versions for debugging
    print("\n=== EXTRACTED OFFICE/SHAREPOINT/EXCHANGE VERSIONS ===")
    for ip in sorted(ip_office_product_map.keys()):
        print(f"IP {ip}: Office = {ip_office_product_map[ip]}")
    for ip in sorted(ip_sharepoint_product_map.keys()):
        print(f"IP {ip}: SharePoint = {ip_sharepoint_product_map[ip]}")
    for ip in sorted(ip_exchange_product_map.keys()):
        print(f"IP {ip}: Exchange = {ip_exchange_product_map[ip]}")
    print("=============================================\n")

    return solutions, ip_product_map, ip_os_family_map, ip_hostname_map, ip_office_product_map, ip_sharepoint_product_map, ip_exchange_product_map

def read_and_sort_csv(csv_file):
    print("Reading and sorting CSV file...")
    df = pd.read_csv(csv_file)
    df_sorted = df.sort_values(by="Asset IP Address")
    print(f"CSV file loaded with {len(df)} records.")
    return df_sorted

def process_vulnerabilities(df_sorted, solutions, ip_product_map, ip_os_family_map, ip_hostname_map, ip_office_product_map, ip_sharepoint_product_map, ip_exchange_product_map):
    print("Processing vulnerabilities...")
    output_data_main = []
    output_data_windows = []

    processed_vuln_ids = set()
    processed_solutions_by_ip = {}  # Dictionary to track unique solutions by IP address
    
    # First pass: collect all Chrome, Firefox, Apache HTTPD, and JRE vulnerabilities per IP
    consolidated_vulns_by_ip = {}  # Dictionary to track Chrome/Firefox/Apache HTTPD/JRE vulnerabilities per IP
    
    for _, row in df_sorted.iterrows():
        ip_address = row['Asset IP Address']
        vuln_id = row['Vulnerability ID']
        
        if ip_address not in consolidated_vulns_by_ip:
            consolidated_vulns_by_ip[ip_address] = {"chrome": [], "firefox": [], "apache": [], "jre": [], "edge": [], "office_obsolete": []}
            
        if vuln_id.startswith("mfsa"):
            consolidated_vulns_by_ip[ip_address]["firefox"].append(vuln_id)
        elif vuln_id.startswith("google-chrome"):
            consolidated_vulns_by_ip[ip_address]["chrome"].append(vuln_id)
        elif vuln_id.startswith("apache-httpd"):
            consolidated_vulns_by_ip[ip_address]["apache"].append(vuln_id)
        elif vuln_id.startswith("jre"):
            consolidated_vulns_by_ip[ip_address]["jre"].append(vuln_id)
        elif vuln_id.startswith("microsoft-edge"):
            consolidated_vulns_by_ip[ip_address]["edge"].append(vuln_id)
        elif vuln_id == "microsoft-office-obsolete":
            consolidated_vulns_by_ip[ip_address]["office_obsolete"].append(vuln_id)

    # Second pass: process vulnerabilities with consolidated Chrome/Firefox entries
    for _, row in tqdm(df_sorted.iterrows(), total=df_sorted.shape[0], desc="Processing Rows"):
        ip_address = row['Asset IP Address']
        vuln_id = row['Vulnerability ID']
        
        if (ip_address, vuln_id) in processed_vuln_ids:
            continue

        solution_items = solutions.get(vuln_id, [])
        os_family = ip_os_family_map.get(ip_address, "")
        product_name = ip_product_map.get(ip_address, "")
        hostname = ip_hostname_map.get(ip_address, "")
        office_product = ip_office_product_map.get(ip_address, "")
        sharepoint_product = ip_sharepoint_product_map.get(ip_address, "")
        exchange_product = ip_exchange_product_map.get(ip_address, "")
        # Construct the full IP address string with product name and hostname
        ip_with_details = ip_address
        if product_name:
            ip_with_details += f" - {product_name}"
        if hostname:
            ip_with_details += f" - {hostname}"

        if not solution_items:
            print(f"No solutions found for Vulnerability ID: {vuln_id} (OS Family: {os_family})")
        
        # Initialize the set for the current IP if not done already
        if ip_address not in processed_solutions_by_ip:
            processed_solutions_by_ip[ip_address] = set()

        # Consolidated checks handled inside Windows section

        if os_family.lower() == "windows":
            # Windows-only Office/SharePoint version gating
            vuln_id_lower = vuln_id.lower()
            # Consolidated software entries (only for Windows)
            if vuln_id.startswith("mfsa") and consolidated_vulns_by_ip[ip_address]["firefox"]:
                if "Outdated Firefox" not in processed_solutions_by_ip[ip_address]:
                    firefox_solution = "Outdated Firefox => Install lastest version of Mozilla Firefox from the http://www.mozilla.org/products/firefox/"
                    output_data_windows.append([ip_with_details, firefox_solution])
                    processed_solutions_by_ip[ip_address].add("Outdated Firefox")
                    for firefox_vuln in consolidated_vulns_by_ip[ip_address]["firefox"]:
                        processed_vuln_ids.add((ip_address, firefox_vuln))
                continue

            if vuln_id.startswith("google-chrome") and consolidated_vulns_by_ip[ip_address]["chrome"]:
                if "Outdated Chrome" not in processed_solutions_by_ip[ip_address]:
                    chrome_solution = "Outdated Chrome => Install latest version of Google Chrome from the http://www.google.com/chrome/"
                    output_data_windows.append([ip_with_details, chrome_solution])
                    processed_solutions_by_ip[ip_address].add("Outdated Chrome")
                    for chrome_vuln in consolidated_vulns_by_ip[ip_address]["chrome"]:
                        processed_vuln_ids.add((ip_address, chrome_vuln))
                continue

            if vuln_id.startswith("microsoft-edge") and consolidated_vulns_by_ip[ip_address]["edge"]:
                if "Outdated Microsoft Edge" not in processed_solutions_by_ip[ip_address]:
                    edge_solution = "Outdated Microsoft Edge => Upgrade to the latest version of Microsoft Edge https://www.microsoft.com/en-us/edgeedge://settings/help"
                    output_data_windows.append([ip_with_details, edge_solution])
                    processed_solutions_by_ip[ip_address].add("Outdated Microsoft Edge")
                    for edge_vuln in consolidated_vulns_by_ip[ip_address]["edge"]:
                        processed_vuln_ids.add((ip_address, edge_vuln))
                continue

            if vuln_id.startswith("apache-httpd") and consolidated_vulns_by_ip[ip_address]["apache"]:
                if "Outdated Apache HTTPD" not in processed_solutions_by_ip[ip_address]:
                    apache_solution = "Outdated Apache HTTPD => Download and apply the upgrade from : https://httpd.apache.org/download.cgi"
                    output_data_windows.append([ip_with_details, apache_solution])
                    processed_solutions_by_ip[ip_address].add("Outdated Apache HTTPD")
                    for apache_vuln in consolidated_vulns_by_ip[ip_address]["apache"]:
                        processed_vuln_ids.add((ip_address, apache_vuln))
                continue

            if vuln_id.startswith("jre") and consolidated_vulns_by_ip[ip_address]["jre"]:
                if "Outdated JRE" not in processed_solutions_by_ip[ip_address]:
                    jre_solution = "Outdated JRE => Download and apply the upgrade from:https://www.java.com/en/download/manual.jsp"
                    output_data_windows.append([ip_with_details, jre_solution])
                    processed_solutions_by_ip[ip_address].add("Outdated JRE")
                    for jre_vuln in consolidated_vulns_by_ip[ip_address]["jre"]:
                        processed_vuln_ids.add((ip_address, jre_vuln))
                continue

            if vuln_id == "microsoft-office-obsolete" and consolidated_vulns_by_ip[ip_address]["office_obsolete"]:
                if "Outdated Office" not in processed_solutions_by_ip[ip_address]:
                    office_obsolete_solution = "Outdated Office => Update to supported Office version"
                    output_data_windows.append([ip_with_details, office_obsolete_solution])
                    processed_solutions_by_ip[ip_address].add("Outdated Office")
                    for office_obsolete_vuln in consolidated_vulns_by_ip[ip_address]["office_obsolete"]:
                        processed_vuln_ids.add((ip_address, office_obsolete_vuln))
                continue
            
            # Check for Exchange Server vulnerabilities (case-insensitive)
            if "exchange" in vuln_id_lower and exchange_product:
                exchange_version = exchange_product.lower()
                filtered_exchange_solutions = []
                for solution in solution_items:
                    if exchange_version in solution.lower(): 
                        filtered_exchange_solutions.append(solution)
                
                # Only process if we found Exchange-related solutions
                if filtered_exchange_solutions:
                    # Replace solution_items with filtered solutions
                    solution_items = filtered_exchange_solutions
            
            # Check for windows-hotfix vulnerabilities that may contain Office version information
            if vuln_id.startswith("windows-hotfix") and office_product:
                
                # Filter solutions to only keep those containing the detected Office version
                office_version = office_product.lower()
                filtered_hotfix_solutions = []
                for solution in solution_items:
                    if office_version in solution.lower():
                        filtered_hotfix_solutions.append(solution)
                
                # Only process if we found Office-related solutions
                if filtered_hotfix_solutions:
                    # Replace solution_items with filtered solutions
                        solution_items = filtered_hotfix_solutions
            
            # Check for Office vulnerabilities (case-insensitive) - excluding microsoft-office-obsolete (handled by consolidation)
            if "office" in vuln_id_lower and office_product and vuln_id != "microsoft-office-obsolete":
                
                # Filter solutions to only keep those containing the detected Office version
                office_version = office_product.lower()
                filtered_office_solutions = []
                for solution in solution_items:
                    if office_version in solution.lower():
                        filtered_office_solutions.append(solution)
                
                # Replace solution_items with filtered solutions
                solution_items = filtered_office_solutions
                
                        # Check for SharePoint vulnerabilities (case-insensitive)
            if "sharepoint" in vuln_id_lower and sharepoint_product:
                
                # Filter solutions to only keep those containing the detected SharePoint version
                sharepoint_version = sharepoint_product.lower()
                filtered_sharepoint_solutions = []
                for solution in solution_items:
                    if sharepoint_version in solution.lower():
                        filtered_sharepoint_solutions.append(solution)
                
                # Replace solution_items with filtered solutions
                solution_items = filtered_sharepoint_solutions

            if vuln_id.startswith("msft") or vuln_id.startswith("microsoft-windows"):
                filtered_solutions = [sol for sol in solution_items if product_name in sol]
            else:
                filtered_solutions = solution_items
            for solution in filtered_solutions:
                # Only append unique solutions
                if solution not in processed_solutions_by_ip[ip_address]:
                    output_data_windows.append([ip_with_details, solution])
                    processed_solutions_by_ip[ip_address].add(solution)
            
        else:
            for solution in solution_items:
                # Only append unique solutions
                if solution not in processed_solutions_by_ip[ip_address]:
                    output_data_main.append([ip_with_details, solution])
                    processed_solutions_by_ip[ip_address].add(solution)

        processed_vuln_ids.add((ip_address, vuln_id))

    return output_data_main, output_data_windows

def create_dataframes(output_data_main, output_data_windows):
    print("Converting processed data to DataFrames...")
    df_main = pd.DataFrame(output_data_main, columns=['Asset IP Address', 'Solution'])
    df_windows = pd.DataFrame(output_data_windows, columns=['Asset IP Address', 'Solution'])
    return df_main, df_windows

def process_linux_dataframe(df_main):
    print("Processing solutions for Linux...")
    df_main['Services'] = df_main['Solution'].apply(lambda x: x.split('=>')[0].strip())
    df_main['Solution Details'] = df_main['Solution'].apply(lambda x: '=>'.join(x.split('=>')[1:]).strip())
    df_main.drop(columns=['Solution'], inplace=True)

    df_main = df_main.drop_duplicates(subset=['Asset IP Address', 'Services']).copy()
    df_main.reset_index(drop=True, inplace=True)
    df_main['Owner'] = ' '

    return df_main

def process_windows_dataframe(df_windows):
    print("Splitting solutions for Windows...")
    df_windows['Services'] = df_windows['Solution'].apply(lambda x: '=>'.join(x.split('=>')[:-1]).strip())
    df_windows['Solution Details'] = df_windows['Solution'].apply(lambda x: x.split('=>')[-1].strip())
    df_windows.drop(columns=['Solution'], inplace=True)

    df_windows.reset_index(drop=True, inplace=True)
    df_windows['Owner'] = ' '

    return df_windows

def save_to_excel(df_main, df_windows, output_file):
    print("Writing data to Excel file...")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_main.to_excel(writer, index=False, sheet_name='Linux')
        df_windows.to_excel(writer, index=False, sheet_name='Windows')
    
    print("Setting column widths, row heights, and formatting...")
    wb = load_workbook(output_file)
    linux_sheet = wb['Linux']
    windows_sheet = wb['Windows']

    # Set the column widths
    for col, width in zip(['A', 'B', 'C', 'D'], [30, 80, 100, 80]):
        linux_sheet.column_dimensions[col].width = width
        windows_sheet.column_dimensions[col].width = width

    # Set all row heights to 15
    for sheet in [linux_sheet, windows_sheet]:
        for row in sheet.iter_rows():
            sheet.row_dimensions[row[0].row].height = 15

    def merge_cells(sheet, col):
        current_value = None
        start_row = 2
        
        for row in range(2, sheet.max_row + 1):
            cell = sheet[f'{col}{row}']
            if cell.value != current_value:
                if start_row < row - 1:
                    sheet.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                current_value = cell.value
                start_row = row
        
        if start_row < sheet.max_row:
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=sheet.max_row, end_column=1)

    def set_alignment(sheet):
        for cell in sheet['A']:
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        for cell in sheet['B']:
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

    merge_cells(linux_sheet, 'A')
    merge_cells(windows_sheet, 'A')

    set_alignment(linux_sheet)
    set_alignment(windows_sheet)
    
    wb.save(output_file)
    print(f"Output successfully written to {output_file}")


def gen_solution_report(csv_file, xml_file):
    print("Starting script...")
    
    df_sorted = read_and_sort_csv(csv_file)
    unique_vuln_ids = df_sorted['Vulnerability ID'].unique()
    solutions, ip_product_map, ip_os_family_map, ip_hostname_map, ip_office_product_map, ip_sharepoint_product_map, ip_exchange_product_map = parse_xml_for_solutions_and_products(xml_file, unique_vuln_ids)
    
    output_data_main, output_data_windows = process_vulnerabilities(df_sorted, solutions, ip_product_map, ip_os_family_map, ip_hostname_map, ip_office_product_map, ip_sharepoint_product_map, ip_exchange_product_map)
    df_main, df_windows = create_dataframes(output_data_main, output_data_windows)

    df_main = process_linux_dataframe(df_main)
    df_windows = process_windows_dataframe(df_windows)

    base_name = os.path.splitext(os.path.basename(csv_file))[0]
    output_file = DOWNLOAD_PATH + os.path.join(f"{base_name}_Solution.xlsx")
    print(output_file)
    save_to_excel(df_main, df_windows, output_file)

gen_solution_report("ServerFarm Windows.csv", "ServerFarm Windows.xml")
#gen_solution_report("UAT60.csv", "UAT60.xml")