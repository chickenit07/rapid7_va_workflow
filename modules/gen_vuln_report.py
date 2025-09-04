import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import argparse
import os
import logging
from tqdm import tqdm
from dotenv import load_dotenv

load_dotenv()

# Report Directory
DOWNLOAD_PATH = os.getenv('DOWNLOAD_PATH')
# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def parse_os_info(xml_file):
    """Parse the XML file to extract OS information based on the 'address' attribute."""
    logging.debug('Parsing XML for OS Info...')
    tree = ET.parse(xml_file)
    root = tree.getroot()
    os_info = {}
    
    for node in root.find('nodes').findall('node'):
        address = node.attrib.get('address')
        if address:
            os_elem = node.find('fingerprints/os')
            if os_elem is not None:
                os_family = os_elem.attrib.get('family', 'Unknown')
                os_info[address] = os_family
    
    logging.debug('OS Info parsing complete.')
    return os_info

def parse_xml(xml_file):
    logging.debug('Parsing XML for vulnerabilities and solutions...')
    tree = ET.parse(xml_file)
    root = tree.getroot()
    vulnerabilities = {}
    
    for vuln in root.find('VulnerabilityDefinitions'):
        vuln_id = vuln.attrib.get('id')
        if not vuln_id:
            continue
        
        # Extract description
        description_elem = vuln.find('description/ContainerBlockElement')
        if description_elem is not None:
            description = get_text_from_container(description_elem)
        else:
            # Fallback if no ContainerBlockElement is found
            description_elem = vuln.find('description/ContainerBlockElement/Paragraph')
            description = description_elem.text if description_elem is not None else ""
        
        # Extract solution
        solution = parse_solution(vuln.find('solution'))
        
        vulnerabilities[vuln_id] = {
            'description': description,
            'solution': solution
        }
    
    logging.debug('Vulnerability parsing complete.')
    return vulnerabilities

def get_text_from_container(container_elem):
    texts = []
    if container_elem is not None:
        # Direct text in ContainerBlockElement
        if container_elem.text:
            texts.append(container_elem.text.strip())
        
        # Process child elements
        for elem in container_elem:
            if elem.tag == 'Paragraph':
                texts.append(get_paragraph_text(elem))
            # Extend this to handle other types of elements if needed
    return " ".join(texts)  # Concatenate texts without "=>"

def get_paragraph_text(paragraph_elem):
    if paragraph_elem is None:
        return ""
    
    if paragraph_elem.tag == 'Paragraph':
        # Extract text from current Paragraph
        text = paragraph_elem.text.strip() if paragraph_elem.text else ""
        
        # Extract text from nested Paragraph elements
        for sub_elem in paragraph_elem:
            if sub_elem.tag == 'Paragraph':
                text += " => " + get_paragraph_text(sub_elem)
        
        # Handle URLLink
        url_elem = paragraph_elem.find('URLLink')
        if url_elem is not None:
            url = url_elem.attrib.get('LinkURL', '')
            text += f" {url}"
        
        return text
    
    return ""

def parse_solution(solution_elem):
    if solution_elem is None:
        return ""
    
    solution_texts = []
    
    # Process ContainerBlockElement
    container_elem = solution_elem.find('ContainerBlockElement')
    if container_elem is not None:
        for child in container_elem:
            if child.tag == 'Paragraph':
                paragraphs = [get_paragraph_text(child)]
                solution_texts.append(" => ".join(paragraphs))
            elif child.tag == 'UnorderedList':
                list_items = []
                for item in child.findall('ListItem'):
                    list_item_texts = []
                    for p in item.findall('Paragraph'):
                        list_item_texts.append(get_paragraph_text(p))
                    combined_text = " => ".join(list_item_texts)
                    list_items.append(combined_text)
                
                # Separate list items with "-------"
                solution_texts.append('\n-------\n'.join(list_items))
                    
    return "\n".join(solution_texts)

def generate_report(csv_file, xml_file, output_file):
    logging.debug('Starting report generation...')
    logging.debug('Reading and sorting CSV file...')
    
    # Read the CSV file
    df = pd.read_csv(csv_file)
    logging.info(f'CSV file loaded with {len(df)} records.')
    
    # Parse the XML file for vulnerabilities and OS information
    vulnerabilities = parse_xml(xml_file)
    os_info = parse_os_info(xml_file)
    
    # List to hold output data
    output_data = []
    
    # Fill the output data list with progress bar
    logging.debug('Processing rows...')
    for _, row in tqdm(df.iterrows(), total=len(df), desc="Processing rows"):
        vuln_id = row['Vulnerability ID']
        ip_address = row['Asset IP Address']
        if vuln_id in vulnerabilities:
            output_data.append({
                'Asset IP Address': ip_address,
                'Vulnerability ID': row['Vulnerability ID'],
                'Vulnerability Severity Level': row['Vulnerability Severity Level'],
                'Vulnerability Title': row['Vulnerability Title'],
                'Description': vulnerabilities[vuln_id]['description'],
                'Solution': vulnerabilities[vuln_id]['solution'],
                'Operating System': os_info.get(ip_address, 'Unknown')  # Add OS info or 'Unknown' if not found
            })
    
    # Create a DataFrame from the output data list
    output_df = pd.DataFrame(output_data, columns=[
        'Operating System',  # New column for OS
        'Asset IP Address', 
        'Vulnerability ID', 
        'Vulnerability Severity Level', 
        'Vulnerability Title',
        'Description',
        'Solution'
    ])
    
    # Sort the DataFrame first by 'Operating System' and then by 'Asset IP Address'
    output_df = output_df.sort_values(by=['Operating System', 'Asset IP Address'])
    
    logging.debug('Writing output to Excel file...')
    
    # Write the output DataFrame to an XLSX file with formatting
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        output_df.to_excel(writer, index=False, sheet_name='Report')
        workbook = writer.book
        worksheet = writer.sheets['Report']
        
        # Set column widths
        column_widths = {
            'A': 15,
            'B': 15,  # Adjust for new column
            'C': 25,
            'D': 15,
            'E': 50,
            'F': 80,
            'G': 80
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Set text wrap and center alignment for all cells
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        
        # Set row height
        for row in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 35
        
        # Merge cells with the same 'Operating System' and 'Asset IP Address' and center the text
        def merge_and_center(column_index):
            start_row = None
            for row in range(2, worksheet.max_row + 1):  # Start from 2 to skip header
                cell_value = worksheet.cell(row=row, column=column_index).value
                prev_cell_value = worksheet.cell(row=row-1, column=column_index).value
                
                if cell_value == prev_cell_value:
                    if start_row is None:
                        start_row = row - 1
                else:
                    if start_row is not None:
                        worksheet.merge_cells(start_row=start_row, start_column=column_index, end_row=row-1, end_column=column_index)
                        worksheet.cell(row=start_row, column=column_index).alignment = Alignment(horizontal='center', vertical='center')
                        start_row = None
            
            # Handle the last group of cells
            if start_row is not None:
                worksheet.merge_cells(start_row=start_row, start_column=column_index, end_row=worksheet.max_row, end_column=column_index)
                worksheet.cell(row=start_row, column=column_index).alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge and center cells in the 'Operating System' and 'Asset IP Address' columns
        merge_and_center(1)  # Operating System column
        merge_and_center(2)  # Asset IP Address column
    
    logging.info(f'Report successfully generated: {output_file}')

def gen_vuln_report(csv_file,xml_file):
    
    # Determine the output file name and path
    base_name = os.path.splitext(os.path.basename(csv_file))[0]
    output_file = DOWNLOAD_PATH + os.path.join(f"{base_name}_Vuln.xlsx")
    
    logging.info('Starting script...')
    generate_report(csv_file, xml_file, output_file)

#gen_vuln_report("ServerFarm Windows.csv", "ServerFarm Windows.xml")