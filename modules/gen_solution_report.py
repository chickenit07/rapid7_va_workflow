import pandas as pd
import lxml.etree as ET
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from tqdm import tqdm
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Report Directory
DOWNLOAD_PATH = os.getenv('DOWNLOAD_PATH', '')

def parse_xml_for_solutions_and_products(xml_file, vuln_ids):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    solutions = {}
    ip_product_map = {}
    ip_os_family_map = {}
    ip_hostname_map = {}  # New map for storing hostnames
    ip_office_product_map = {}  # Store Office product per IP
    ip_sharepoint_product_map = {}  # Store SharePoint product per IP
    ip_exchange_product_map = {}  # Store Exchange Server product per IP
    ip_risk_map = {}

    # Extract product names, OS family, and hostnames for each IP address
    for node in root.xpath(".//node"):
        ip_address = node.get("address")
        fingerprint = node.find(".//fingerprints/os")
        hostname_elem = node.find(".//names/name")  # Extract the hostname

        if fingerprint is not None:
            product_name = fingerprint.get("product")
            os_family = fingerprint.get("family")  # Extract the OS family
            ip_product_map[ip_address] = product_name
            ip_os_family_map[ip_address] = os_family

        if hostname_elem is not None:
            hostname = hostname_elem.text.strip()
            ip_hostname_map[ip_address] = hostname

        # Extract risk-score from the node
        risk_score = node.get("risk-score")
        if risk_score:
            try:
                # Keep the risk score as a float to preserve decimal format
                ip_risk_map[ip_address] = float(risk_score)
            except (ValueError, TypeError):
                pass

        # Extract Office and SharePoint products from software fingerprints
        for sw_fp in node.findall('.//software/fingerprint'):
            product_value = sw_fp.get('product', '')
            if product_value:
                # Only capture Office entries that contain a year (like "Office 2013", "Office 2016")
                if 'Office' in product_value and any(char.isdigit() for char in product_value) and ip_address not in ip_office_product_map:
                    ip_office_product_map[ip_address] = product_value
                # Only capture SharePoint entries that contain a year (like "SharePoint 2016", "SharePoint 2019")
                if 'SharePoint' in product_value and any(char.isdigit() for char in product_value) and ip_address not in ip_sharepoint_product_map:
                    ip_sharepoint_product_map[ip_address] = product_value
                # Only capture Exchange Server entries that contain a year (like "Exchange Server 2016", "Exchange Server 2019")
                if 'Exchange Server' in product_value and any(char.isdigit() for char in product_value) and ip_address not in ip_exchange_product_map:
                    ip_exchange_product_map[ip_address] = product_value

    for vuln_id in tqdm(vuln_ids, desc="Parsing XML for Solutions"):
        solution_list = []

        # Check for solutions under ListItem
        xpath_query_list_item = f".//vulnerability[@id='{vuln_id}']/solution/ContainerBlockElement/UnorderedList/ListItem"
        list_items = root.xpath(xpath_query_list_item)

        for item in list_items:
            paragraphs = item.xpath(".//Paragraph")
            item_texts = []

            for i, para in enumerate(paragraphs):
                text_parts = []
                for elem in para.iter():
                    if elem.tag == "URLLink":
                        text_parts.append(elem.get("LinkURL"))
                    elif elem.text:
                        text_parts.append(elem.text.strip())
                paragraph_text = "".join(text_parts)

                if i == 0:
                    item_texts.append(paragraph_text)
                else:
                    item_texts.append(f"=> {paragraph_text}")

            # Combine all paragraphs into one solution instead of separate solutions
            full_item_text = " ".join(item_texts)
            if full_item_text.strip():  # Only add if there's actual content
                solution_list.append(full_item_text)

        # Check for solutions directly under ContainerBlockElement/Paragraph
        xpath_query_paragraph = f".//vulnerability[@id='{vuln_id}']/solution/ContainerBlockElement/Paragraph"
        paragraphs = root.xpath(xpath_query_paragraph)

        # Combine related paragraphs into single solutions instead of separate ones
        if len(paragraphs) > 1:
            # Multiple paragraphs - combine them into one solution
            combined_texts = []
            for para in paragraphs:
                text_parts = []
                for elem in para.iter():
                    if elem.tag == "URLLink":
                        text_parts.append(elem.get("LinkURL"))
                    elif elem.text:
                        text_parts.append(elem.text.strip())
                paragraph_text = "".join(text_parts)
                if paragraph_text:
                    combined_texts.append(paragraph_text)
            
            if combined_texts:
                # Join with " => " to create a proper solution format
                combined_solution = " => ".join(combined_texts)
                solution_list.append(combined_solution)
        else:
            # Single paragraph - handle as before
            for para in paragraphs:
                text_parts = []
                for elem in para.iter():
                    if elem.tag == "URLLink":
                        text_parts.append(elem.get("LinkURL"))
                    elif elem.text:
                        text_parts.append(elem.text.strip())
                paragraph_text = "".join(text_parts)
                if paragraph_text:
                    solution_list.append(f"{vuln_id} => {paragraph_text}")

        # Check for solutions nested within ContainerBlockElement/Paragraph/Paragraph
        xpath_query_nested_paragraph = f".//vulnerability[@id='{vuln_id}']/solution/ContainerBlockElement/Paragraph/Paragraph"
        nested_paragraphs = root.xpath(xpath_query_nested_paragraph)

        for para in nested_paragraphs:
            text_parts = []
            for elem in para.iter():
                if elem.tag == "URLLink":
                    text_parts.append(elem.get("LinkURL"))
                elif elem.text:
                    text_parts.append(elem.text.strip())
            nested_text = "".join(text_parts)
            if nested_text:
                solution_list.append(f"{vuln_id} => {nested_text}")

        solutions[vuln_id] = solution_list

    return solutions, ip_product_map, ip_os_family_map, ip_hostname_map, ip_office_product_map, ip_sharepoint_product_map, ip_exchange_product_map, ip_risk_map

def read_and_sort_csv(csv_file):
    print("Reading and sorting CSV file...")
    df = pd.read_csv(csv_file)
    df_sorted = df.sort_values(by="Asset IP Address")
    print(f"CSV file loaded with {len(df)} records.")
    return df_sorted

def process_vulnerabilities(df_sorted, solutions, ip_product_map, ip_os_family_map, ip_hostname_map, ip_office_product_map, ip_sharepoint_product_map, ip_exchange_product_map, ip_risk_map):
    print("Processing vulnerabilities...")
    output_data_main = []
    output_data_windows = []

    processed_vuln_ids = set()
    processed_solutions_by_ip = {}  # Dictionary to track unique solutions by IP address
    
    # Define consolidated vulnerability types and their patterns
    consolidated_vuln_types = {
        "firefox": {"prefix": "mfsa", "solution": "Outdated Firefox => Install lastest version of Mozilla Firefox from the http://www.mozilla.org/products/firefox/", "key": "Outdated Firefox"},
        "chrome": {"prefix": "google-chrome", "solution": "Outdated Chrome => Install latest version of Google Chrome from the http://www.google.com/chrome/", "key": "Outdated Chrome"},
        "edge": {"prefix": "microsoft-edge", "solution": "Outdated Microsoft Edge => Upgrade to the latest version of Microsoft Edge https://www.microsoft.com/en-us/edgeedge://settings/help", "key": "Outdated Microsoft Edge"},
        "apache": {"prefix": "apache-httpd", "solution": "Outdated Apache HTTPD => Download and apply the upgrade from : https://httpd.apache.org/download.cgi", "key": "Outdated Apache HTTPD"},
        "jre": {"prefix": "jre", "solution": "Outdated JRE => Download and apply the upgrade from:https://www.java.com/en/download/manual.jsp", "key": "Outdated JRE"},
        "office_obsolete": {"prefix": "microsoft-office-obsolete", "solution": "Outdated Office => Update to supported Office version", "key": "Outdated Office", "exact_match": True},
        "tomcat": {"prefix": "apache-tomcat", "solution": "Outdated Apache Tomcat => Download and apply the upgrade appropriated version from: http://archive.apache.org/dist/tomcat/", "key": "Outdated Apache Tomcat"},
        "winrar": {"prefix": "rarlab-winrar", "solution": "Outdated Winrar => Upgrade Rarlab WinRAR to version latest from https://www.win-rar.com/download.html", "key": "Outdated Winrar"},
        "openssl": {"prefix": "http-openssl", "solution": "Outdated/Insecure OpenSSL => Upgrade to latest version:https://openssl-library.org/source/", "key": "Outdated/Insecure OpenSSL"}
    }
    
    # First pass: collect all consolidated vulnerabilities per IP
    consolidated_vulns_by_ip = {}  # Dictionary to track consolidated vulnerabilities per IP
    
    # Helper function to handle consolidated vulnerability checks
    def handle_consolidated_vuln(vuln_type, vuln_id, ip_address, consolidated_vulns_by_ip, processed_solutions_by_ip, processed_vuln_ids, output_data_windows, ip_with_details):
        """Handle consolidated vulnerability processing"""
        config = consolidated_vuln_types[vuln_type]
        key = config["key"]
        
        if key not in processed_solutions_by_ip[ip_address]:
            output_data_windows.append([ip_with_details, config["solution"]])
            processed_solutions_by_ip[ip_address].add(key)
            for vuln in consolidated_vulns_by_ip[ip_address][vuln_type]:
                processed_vuln_ids.add((ip_address, vuln))
            return True  # Indicates this vulnerability was handled
        return False  # Indicates this vulnerability was already processed
    
    for _, row in df_sorted.iterrows():
        ip_address = row['Asset IP Address']
        vuln_id = row['Vulnerability ID']
        
        if ip_address not in consolidated_vulns_by_ip:
            consolidated_vulns_by_ip[ip_address] = {vuln_type: [] for vuln_type in consolidated_vuln_types.keys()}
            
        # Check each vulnerability type
        for vuln_type, config in consolidated_vuln_types.items():
            if config.get("exact_match", False):
                if vuln_id == config["prefix"]:
                    consolidated_vulns_by_ip[ip_address][vuln_type].append(vuln_id)
            else:
                if vuln_id.startswith(config["prefix"]):
                    consolidated_vulns_by_ip[ip_address][vuln_type].append(vuln_id)

    # Second pass: process vulnerabilities with consolidated checks
    for _, row in tqdm(df_sorted.iterrows(), total=df_sorted.shape[0], desc="Processing Rows"):
        ip_address = row['Asset IP Address']
        vuln_id = row['Vulnerability ID']
        
        if (ip_address, vuln_id) in processed_vuln_ids:
            continue

        solution_items = solutions.get(vuln_id, [])
        os_family = ip_os_family_map.get(ip_address, "")
        product_name = ip_product_map.get(ip_address, "")
        hostname = ip_hostname_map.get(ip_address, "")
        office_product = ip_office_product_map.get(ip_address, "")
        sharepoint_product = ip_sharepoint_product_map.get(ip_address, "")
        exchange_product = ip_exchange_product_map.get(ip_address, "")
        risk_score = ip_risk_map.get(ip_address, "")
        # Construct the full IP address string with product name, hostname, and risk score
        ip_with_details = ip_address
        if product_name:
            ip_with_details += f" - {product_name}"
        if hostname:
            ip_with_details += f" - {hostname}"
        if risk_score:
            # Format risk score with thousands separator, e.g., 666,634
            try:
                formatted_risk = f"{float(risk_score):,.0f}"
            except Exception:
                formatted_risk = str(risk_score)
            ip_with_details += f" (Risk: {formatted_risk})"

        if not solution_items:
            print(f"No solutions found for Vulnerability ID: {vuln_id} (OS Family: {os_family})")
        
        # Initialize the set for the current IP if not done already
        if ip_address not in processed_solutions_by_ip:
            processed_solutions_by_ip[ip_address] = set()

        # Consolidated checks handled inside Windows section

        if os_family.lower() == "windows":
            vuln_id_lower = vuln_id.lower()
            
            # Check for consolidated vulnerabilities using the helper function
            handled = False
            for vuln_type in consolidated_vuln_types.keys():
                if vuln_type == "office_obsolete":
                    if vuln_id == consolidated_vuln_types[vuln_type]["prefix"] and consolidated_vulns_by_ip[ip_address][vuln_type]:
                        handled = handle_consolidated_vuln(vuln_type, vuln_id, ip_address, consolidated_vulns_by_ip, processed_solutions_by_ip, processed_vuln_ids, output_data_windows, ip_with_details)
                        if handled:
                            break
                else:
                    if vuln_id.startswith(consolidated_vuln_types[vuln_type]["prefix"]) and consolidated_vulns_by_ip[ip_address][vuln_type]:
                        handled = handle_consolidated_vuln(vuln_type, vuln_id, ip_address, consolidated_vulns_by_ip, processed_solutions_by_ip, processed_vuln_ids, output_data_windows, ip_with_details)
                        if handled:
                            break
            
            if handled:
                continue
            # Check for Exchange Server vulnerabilities (case-insensitive)
            if "exchange" in vuln_id_lower and exchange_product:
                exchange_version = exchange_product.lower()
                filtered_exchange_solutions = []
                for solution in solution_items:
                    if exchange_version in solution.lower(): 
                        filtered_exchange_solutions.append(solution)
                
                # Only process if we found Exchange-related solutions
                if filtered_exchange_solutions:
                    # Replace solution_items with filtered solutions
                    solution_items = filtered_exchange_solutions
            
            # Check for windows-hotfix vulnerabilities that may contain Office version information
            if vuln_id.startswith("windows-hotfix") and office_product:
                
                # Filter solutions to only keep those containing the detected Office version
                office_version = office_product.lower()
                filtered_hotfix_solutions = []
                for solution in solution_items:
                    if office_version in solution.lower():
                        filtered_hotfix_solutions.append(solution)
                
                # Only process if we found Office-related solutions
                if filtered_hotfix_solutions:
                    # Replace solution_items with filtered solutions
                        solution_items = filtered_hotfix_solutions
                         
            # Check for SharePoint vulnerabilities (case-insensitive)
            if "sharepoint" in vuln_id_lower and sharepoint_product:
                
                # Filter solutions to only keep those containing the detected SharePoint version
                sharepoint_version = sharepoint_product.lower()
                filtered_sharepoint_solutions = []
                for solution in solution_items:
                    if sharepoint_version in solution.lower():
                        filtered_sharepoint_solutions.append(solution)
                
                # Replace solution_items with filtered solutions
                solution_items = filtered_sharepoint_solutions
            
            # Check for Office vulnerabilities (case-insensitive) and filter by detected Office version
            if "office" in vuln_id_lower and office_product and vuln_id != "microsoft-office-obsolete":
                
                # Filter solutions to only keep those containing the detected Office version
                office_version = office_product.lower()
                filtered_office_solutions = []
                for solution in solution_items:
                    if office_version in solution.lower():
                        filtered_office_solutions.append(solution)
                
                # Only process if we found Office-related solutions
                if filtered_office_solutions:
                    # Replace solution_items with filtered solutions
                    solution_items = filtered_office_solutions
                else:
                    # If no Office version matches, skip this vulnerability entirely
                    continue

            # Apply Windows version filtering (keep existing logic)
            if vuln_id.startswith("msft") or vuln_id.startswith("microsoft-windows"):
                filtered_solutions = [sol for sol in solution_items if product_name in sol]
            else:
                filtered_solutions = solution_items
            
            # Apply Office version filtering ONLY to Office-related solutions
            if office_product and filtered_solutions:
                office_version = office_product.lower()
                office_filtered_solutions = []
                for solution in filtered_solutions:
                    # Check if this solution contains Office version information
                    if any(office_keyword in solution.lower() for office_keyword in ['office', 'word', 'excel', 'powerpoint', 'outlook']):
                        # If it's an Office-related solution, only include if it matches the detected version
                        if office_version in solution.lower():
                            office_filtered_solutions.append(solution)
                    else:
                        # If it's not Office-related, include it as-is
                        office_filtered_solutions.append(solution)
                
                # Replace filtered_solutions with Office-filtered results
                filtered_solutions = office_filtered_solutions
            for solution in filtered_solutions:
                # Only append unique solutions
                if solution not in processed_solutions_by_ip[ip_address]:
                    output_data_windows.append([ip_with_details, solution])
                    processed_solutions_by_ip[ip_address].add(solution)
            
        else:
            for solution in solution_items:
                # Only append unique solutions
                if solution not in processed_solutions_by_ip[ip_address]:
                    output_data_main.append([ip_with_details, solution])
                    processed_solutions_by_ip[ip_address].add(solution)

        processed_vuln_ids.add((ip_address, vuln_id))

    return output_data_main, output_data_windows

def create_dataframes(output_data_main, output_data_windows):
    print("Converting processed data to DataFrames...")
    
    # Sort the data by IP address and then by vulnerability ID (extracted from solution)
    def sort_by_ip_and_vuln(data_list):
        # Group by IP address
        ip_groups = {}
        for row in data_list:
            ip = row[0]
            solution = row[1]
            if ip not in ip_groups:
                ip_groups[ip] = []
            ip_groups[ip].append(row)
        
        # Sort vulnerabilities within each IP group
        sorted_data = []
        for ip in sorted(ip_groups.keys()):
            # Extract vulnerability ID from solution for sorting
            ip_rows = ip_groups[ip]
            ip_rows.sort(key=lambda x: x[1].split('=>')[0].strip() if '=>' in x[1] else x[1])
            sorted_data.extend(ip_rows)
        
        return sorted_data
    
    # Sort both main and Windows data
    sorted_main_data = sort_by_ip_and_vuln(output_data_main)
    sorted_windows_data = sort_by_ip_and_vuln(output_data_windows)
    
    df_main = pd.DataFrame(sorted_main_data, columns=['Asset IP Address', 'Solution'])
    df_windows = pd.DataFrame(sorted_windows_data, columns=['Asset IP Address', 'Solution'])
    return df_main, df_windows

def process_linux_dataframe(df_main):
    print("Processing solutions for Linux...")
    df_main['Services'] = df_main['Solution'].apply(lambda x: x.split('=>')[0].strip())
    df_main['Solution Details'] = df_main['Solution'].apply(lambda x: '=>'.join(x.split('=>')[1:]).strip())
    df_main.drop(columns=['Solution'], inplace=True)

    df_main = df_main.drop_duplicates(subset=['Asset IP Address', 'Services']).copy()
    df_main.reset_index(drop=True, inplace=True)
    df_main['Owner'] = ' '

    return df_main

def process_windows_dataframe(df_windows):
    print("Splitting solutions for Windows...")
    df_windows['Services'] = df_windows['Solution'].apply(lambda x: '=>'.join(x.split('=>')[:-1]).strip())
    df_windows['Solution Details'] = df_windows['Solution'].apply(lambda x: x.split('=>')[-1].strip())
    df_windows.drop(columns=['Solution'], inplace=True)

    df_windows.reset_index(drop=True, inplace=True)
    df_windows['Owner'] = ' '

    return df_windows

def save_to_excel(df_main, df_windows, output_file):
    print("Writing data to Excel file...")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_main.to_excel(writer, index=False, sheet_name='Linux')
        df_windows.to_excel(writer, index=False, sheet_name='Windows')
    
    print("Setting column widths, row heights, and formatting...")
    wb = load_workbook(output_file)
    linux_sheet = wb['Linux']
    windows_sheet = wb['Windows']

    # Set the column widths
    for col, width in zip(['A', 'B', 'C', 'D'], [30, 80, 100, 80]):
        linux_sheet.column_dimensions[col].width = width
        windows_sheet.column_dimensions[col].width = width

    # Set all row heights to 15
    for sheet in [linux_sheet, windows_sheet]:
        for row in sheet.iter_rows():
            sheet.row_dimensions[row[0].row].height = 15

    def merge_cells(sheet, col):
        current_value = None
        start_row = 2
        
        for row in range(2, sheet.max_row + 1):
            cell = sheet[f'{col}{row}']
            if cell.value != current_value:
                if start_row < row - 1:
                    sheet.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                current_value = cell.value
                start_row = row
        
        if start_row < sheet.max_row:
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=sheet.max_row, end_column=1)

    def set_alignment(sheet):
        for cell in sheet['A']:
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        for cell in sheet['B']:
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

    merge_cells(linux_sheet, 'A')
    merge_cells(windows_sheet, 'A')

    set_alignment(linux_sheet)
    set_alignment(windows_sheet)
    
    wb.save(output_file)
    print(f"Output successfully written to {output_file}")


def gen_solution_report(csv_file, xml_file):
    print("Starting script...")
    
    df_sorted = read_and_sort_csv(csv_file)
    
    # Get all unique vulnerability IDs from the CSV
    unique_vuln_ids = df_sorted['Vulnerability ID'].unique()
    print(f"Found {len(unique_vuln_ids)} unique vulnerabilities to parse solutions for")
    
    # Parse solutions for all vulnerabilities
    solutions, ip_product_map, ip_os_family_map, ip_hostname_map, ip_office_product_map, ip_sharepoint_product_map, ip_exchange_product_map, ip_risk_map = parse_xml_for_solutions_and_products(xml_file, unique_vuln_ids)
    
    output_data_main, output_data_windows = process_vulnerabilities(df_sorted, solutions, ip_product_map, ip_os_family_map, ip_hostname_map, ip_office_product_map, ip_sharepoint_product_map, ip_exchange_product_map, ip_risk_map)
        
    df_main, df_windows = create_dataframes(output_data_main, output_data_windows)

    df_main = process_linux_dataframe(df_main)
    df_windows = process_windows_dataframe(df_windows)

    base_name = os.path.splitext(os.path.basename(csv_file))[0]
    output_file = DOWNLOAD_PATH + os.path.join(f"{base_name}_Solution.xlsx")
    print(output_file)
    save_to_excel(df_main, df_windows, output_file)

#gen_solution_report("UDNganh.csv", "UDNganh.xml")
#gen_solution_report("ServerFarm Windows.csv", "ServerFarm Windows.xml")
#gen_solution_report("UAT59.csv", "UAT59.xml")
#gen_solution_report("UAT60.csv", "UAT60.xml")