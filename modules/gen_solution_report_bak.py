import pandas as pd
import lxml.etree as ET
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from tqdm import tqdm
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Report Directory
DOWNLOAD_PATH = os.getenv('DOWNLOAD_PATH', '')

def parse_xml_for_solutions_and_products(xml_file, vuln_ids):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    solutions = {}
    ip_product_map = {}
    ip_os_family_map = {}
    ip_hostname_map = {}  # New map for storing hostnames

    # Extract product names, OS family, and hostnames for each IP address
    for node in root.xpath(".//node"):
        ip_address = node.get("address")
        fingerprint = node.find(".//fingerprints/os")
        hostname_elem = node.find(".//names/name")  # Extract the hostname

        if fingerprint is not None:
            product_name = fingerprint.get("product")
            os_family = fingerprint.get("family")  # Extract the OS family
            ip_product_map[ip_address] = product_name
            ip_os_family_map[ip_address] = os_family

        if hostname_elem is not None:
            hostname = hostname_elem.text.strip()
            ip_hostname_map[ip_address] = hostname

    for vuln_id in tqdm(vuln_ids, desc="Parsing XML for Solutions"):
        solution_list = []

        # Check for solutions under ListItem
        xpath_query_list_item = f".//vulnerability[@id='{vuln_id}']/solution/ContainerBlockElement/UnorderedList/ListItem"
        list_items = root.xpath(xpath_query_list_item)

        for item in list_items:
            paragraphs = item.xpath(".//Paragraph")
            item_texts = []

            for i, para in enumerate(paragraphs):
                text_parts = []
                for elem in para.iter():
                    if elem.tag == "URLLink":
                        text_parts.append(elem.get("LinkURL"))
                    elif elem.text:
                        text_parts.append(elem.text.strip())
                paragraph_text = "".join(text_parts)

                if i == 0:
                    item_texts.append(paragraph_text)
                else:
                    item_texts.append(f"=> {paragraph_text}")

            full_item_text = " ".join(item_texts)
            solution_list.append(full_item_text)

        # Check for solutions directly under ContainerBlockElement/Paragraph
        xpath_query_paragraph = f".//vulnerability[@id='{vuln_id}']/solution/ContainerBlockElement/Paragraph"
        paragraphs = root.xpath(xpath_query_paragraph)

        for para in paragraphs:
            text_parts = []
            for elem in para.iter():
                if elem.tag == "URLLink":
                    text_parts.append(elem.get("LinkURL"))
                elif elem.text:
                    text_parts.append(elem.text.strip())
            paragraph_text = "".join(text_parts)
            if paragraph_text:
                solution_list.append(f"{vuln_id} => {paragraph_text}")

        # Check for solutions nested within ContainerBlockElement/Paragraph/Paragraph
        xpath_query_nested_paragraph = f".//vulnerability[@id='{vuln_id}']/solution/ContainerBlockElement/Paragraph/Paragraph"
        nested_paragraphs = root.xpath(xpath_query_nested_paragraph)

        for para in nested_paragraphs:
            text_parts = []
            for elem in para.iter():
                if elem.tag == "URLLink":
                    text_parts.append(elem.get("LinkURL"))
                elif elem.text:
                    text_parts.append(elem.text.strip())
            nested_text = "".join(text_parts)
            if nested_text:
                solution_list.append(f"{vuln_id} => {nested_text}")

        solutions[vuln_id] = solution_list

    return solutions, ip_product_map, ip_os_family_map, ip_hostname_map  # Return the hostname map

def read_and_sort_csv(csv_file):
    print("Reading and sorting CSV file...")
    df = pd.read_csv(csv_file)
    df_sorted = df.sort_values(by="Asset IP Address")
    print(f"CSV file loaded with {len(df)} records.")
    return df_sorted

def process_vulnerabilities(df_sorted, solutions, ip_product_map, ip_os_family_map, ip_hostname_map):
    print("Processing vulnerabilities...")
    output_data_main = []
    output_data_windows = []

    processed_vuln_ids = set()
    processed_solutions_by_ip = {}  # Dictionary to track unique solutions by IP address
    
    # First pass: collect all Chrome, Firefox, Apache HTTPD, and JRE vulnerabilities per IP
    consolidated_vulns_by_ip = {}  # Dictionary to track Chrome/Firefox/Apache HTTPD/JRE vulnerabilities per IP
    
    for _, row in df_sorted.iterrows():
        ip_address = row['Asset IP Address']
        vuln_id = row['Vulnerability ID']
        
        if ip_address not in consolidated_vulns_by_ip:
            consolidated_vulns_by_ip[ip_address] = {"chrome": [], "firefox": [], "apache": [], "jre": []}
            
        if vuln_id.startswith("mfsa"):
            consolidated_vulns_by_ip[ip_address]["firefox"].append(vuln_id)
        elif vuln_id.startswith("google-chrome"):
            consolidated_vulns_by_ip[ip_address]["chrome"].append(vuln_id)
        elif vuln_id.startswith("apache-httpd"):
            consolidated_vulns_by_ip[ip_address]["apache"].append(vuln_id)
        elif vuln_id.startswith("jre"):
            consolidated_vulns_by_ip[ip_address]["jre"].append(vuln_id)

    # Second pass: process vulnerabilities with consolidated Chrome/Firefox entries
    for _, row in tqdm(df_sorted.iterrows(), total=df_sorted.shape[0], desc="Processing Rows"):
        ip_address = row['Asset IP Address']
        vuln_id = row['Vulnerability ID']
        
        if (ip_address, vuln_id) in processed_vuln_ids:
            continue

        solution_items = solutions.get(vuln_id, [])
        os_family = ip_os_family_map.get(ip_address, "")
        product_name = ip_product_map.get(ip_address, "")
        hostname = ip_hostname_map.get(ip_address, "")

        # Construct the full IP address string with product name and hostname
        ip_with_details = ip_address
        if product_name:
            ip_with_details += f" - {product_name}"
        if hostname:
            ip_with_details += f" - {hostname}"

        if not solution_items:
            print(f"No solutions found for Vulnerability ID: {vuln_id} (OS Family: {os_family})")
        
        # Initialize the set for the current IP if not done already
        if ip_address not in processed_solutions_by_ip:
            processed_solutions_by_ip[ip_address] = set()

        # Check for Firefox, Chrome, Apache HTTPD, and JRE vulnerabilities - add consolidated entries first
        if vuln_id.startswith("mfsa") and consolidated_vulns_by_ip[ip_address]["firefox"]:
            # Add consolidated Firefox solution only once
            if "Outdated Firefox" not in processed_solutions_by_ip[ip_address]:
                firefox_solution = "Outdated Firefox => Install lastest version of Mozilla Firefox from the http://www.mozilla.org/products/firefox/"
                if os_family.lower() == "windows":
                    output_data_windows.append([ip_with_details, firefox_solution])
                else:
                    output_data_main.append([ip_with_details, firefox_solution])
                processed_solutions_by_ip[ip_address].add("Outdated Firefox")
                # Mark all Firefox vulnerabilities for this IP as processed
                for firefox_vuln in consolidated_vulns_by_ip[ip_address]["firefox"]:
                    processed_vuln_ids.add((ip_address, firefox_vuln))
            continue
            
        elif vuln_id.startswith("google-chrome") and consolidated_vulns_by_ip[ip_address]["chrome"]:
            # Add consolidated Chrome solution only once
            if "Outdated Chrome" not in processed_solutions_by_ip[ip_address]:
                chrome_solution = "Outdated Chrome => Install latest version of Google Chrome from the http://www.google.com/chrome/"
                if os_family.lower() == "windows":
                    output_data_windows.append([ip_with_details, chrome_solution])
                else:
                    output_data_main.append([ip_with_details, chrome_solution])
                processed_solutions_by_ip[ip_address].add("Outdated Chrome")
                # Mark all Chrome vulnerabilities for this IP as processed
                for chrome_vuln in consolidated_vulns_by_ip[ip_address]["chrome"]:
                    processed_vuln_ids.add((ip_address, chrome_vuln))
            continue
            
        elif vuln_id.startswith("apache-httpd") and consolidated_vulns_by_ip[ip_address]["apache"]:
            # Add consolidated Apache HTTPD solution only once
            if "Outdated Apache HTTPD" not in processed_solutions_by_ip[ip_address]:
                apache_solution = "Outdated Apache HTTPD => Download and apply the upgrade from : https://httpd.apache.org/download.cgi"
                if os_family.lower() == "windows":
                    output_data_windows.append([ip_with_details, apache_solution])
                else:
                    output_data_main.append([ip_with_details, apache_solution])
                processed_solutions_by_ip[ip_address].add("Outdated Apache HTTPD")
                # Mark all Apache HTTPD vulnerabilities for this IP as processed
                for apache_vuln in consolidated_vulns_by_ip[ip_address]["apache"]:
                    processed_vuln_ids.add((ip_address, apache_vuln))
            continue

        elif vuln_id.startswith("jre") and consolidated_vulns_by_ip[ip_address]["jre"]:
            # Add consolidated JRE solution only once
            if "Outdated JRE" not in processed_solutions_by_ip[ip_address]:
                jre_solution = "Outdated JRE => Download and apply the upgrade from:https://www.java.com/en/download/manual.jsp"
                if os_family.lower() == "windows":
                    output_data_windows.append([ip_with_details, jre_solution])
                else:
                    output_data_main.append([ip_with_details, jre_solution])
                processed_solutions_by_ip[ip_address].add("Outdated JRE")
                # Mark all JRE vulnerabilities for this IP as processed
                for jre_vuln in consolidated_vulns_by_ip[ip_address]["jre"]:
                    processed_vuln_ids.add((ip_address, jre_vuln))
            continue

        if os_family.lower() == "windows":
            if vuln_id.startswith("msft") or vuln_id.startswith("microsoft-windows"):
                filtered_solutions = [sol for sol in solution_items if product_name in sol]
            else:
                filtered_solutions = solution_items
            for solution in filtered_solutions:
                # Only append unique solutions
                if solution not in processed_solutions_by_ip[ip_address]:
                    output_data_windows.append([ip_with_details, solution])
                    processed_solutions_by_ip[ip_address].add(solution)
            
        else:
            for solution in solution_items:
                # Only append unique solutions
                if solution not in processed_solutions_by_ip[ip_address]:
                    output_data_main.append([ip_with_details, solution])
                    processed_solutions_by_ip[ip_address].add(solution)

        processed_vuln_ids.add((ip_address, vuln_id))

    return output_data_main, output_data_windows

def create_dataframes(output_data_main, output_data_windows):
    print("Converting processed data to DataFrames...")
    df_main = pd.DataFrame(output_data_main, columns=['Asset IP Address', 'Solution'])
    df_windows = pd.DataFrame(output_data_windows, columns=['Asset IP Address', 'Solution'])
    return df_main, df_windows

def process_linux_dataframe(df_main):
    print("Processing solutions for Linux...")
    df_main['Services'] = df_main['Solution'].apply(lambda x: x.split('=>')[0].strip())
    df_main['Solution Details'] = df_main['Solution'].apply(lambda x: '=>'.join(x.split('=>')[1:]).strip())
    df_main.drop(columns=['Solution'], inplace=True)

    df_main = df_main.drop_duplicates(subset=['Asset IP Address', 'Services']).copy()
    df_main.reset_index(drop=True, inplace=True)
    df_main['Owner'] = ' '

    return df_main

def process_windows_dataframe(df_windows):
    print("Splitting solutions for Windows...")
    df_windows['Services'] = df_windows['Solution'].apply(lambda x: '=>'.join(x.split('=>')[:-1]).strip())
    df_windows['Solution Details'] = df_windows['Solution'].apply(lambda x: x.split('=>')[-1].strip())
    df_windows.drop(columns=['Solution'], inplace=True)

    df_windows.reset_index(drop=True, inplace=True)
    df_windows['Owner'] = ' '

    return df_windows

def save_to_excel(df_main, df_windows, output_file):
    print("Writing data to Excel file...")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_main.to_excel(writer, index=False, sheet_name='Linux')
        df_windows.to_excel(writer, index=False, sheet_name='Windows')
    
    print("Setting column widths, row heights, and formatting...")
    wb = load_workbook(output_file)
    linux_sheet = wb['Linux']
    windows_sheet = wb['Windows']

    # Set the column widths
    for col, width in zip(['A', 'B', 'C', 'D'], [30, 80, 100, 80]):
        linux_sheet.column_dimensions[col].width = width
        windows_sheet.column_dimensions[col].width = width

    # Set all row heights to 15
    for sheet in [linux_sheet, windows_sheet]:
        for row in sheet.iter_rows():
            sheet.row_dimensions[row[0].row].height = 15

    def merge_cells(sheet, col):
        current_value = None
        start_row = 2
        
        for row in range(2, sheet.max_row + 1):
            cell = sheet[f'{col}{row}']
            if cell.value != current_value:
                if start_row < row - 1:
                    sheet.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                current_value = cell.value
                start_row = row
        
        if start_row < sheet.max_row:
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=sheet.max_row, end_column=1)

    def set_alignment(sheet):
        for cell in sheet['A']:
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        for cell in sheet['B']:
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

    merge_cells(linux_sheet, 'A')
    merge_cells(windows_sheet, 'A')

    set_alignment(linux_sheet)
    set_alignment(windows_sheet)
    
    wb.save(output_file)
    print(f"Output successfully written to {output_file}")


def gen_solution_report(csv_file, xml_file):
    print("Starting script...")
    
    df_sorted = read_and_sort_csv(csv_file)
    unique_vuln_ids = df_sorted['Vulnerability ID'].unique()
    solutions, ip_product_map, ip_os_family_map, ip_hostname_map = parse_xml_for_solutions_and_products(xml_file, unique_vuln_ids)
    
    output_data_main, output_data_windows = process_vulnerabilities(df_sorted, solutions, ip_product_map, ip_os_family_map, ip_hostname_map)
    df_main, df_windows = create_dataframes(output_data_main, output_data_windows)

    df_main = process_linux_dataframe(df_main)
    df_windows = process_windows_dataframe(df_windows)

    base_name = os.path.splitext(os.path.basename(csv_file))[0]
    output_file = DOWNLOAD_PATH + os.path.join(f"{base_name}_Solution.xlsx")
    print(output_file)
    save_to_excel(df_main, df_windows, output_file)

gen_solution_report("ServerFarm Windows.csv", "ServerFarm Windows.xml")
